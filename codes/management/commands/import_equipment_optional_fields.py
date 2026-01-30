import re
from collections import defaultdict

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from codes.models import (
    DropdownList,
    DropdownOption,
    AssetField,
    EquipmentOptionalField,
)


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip()).lower()


def _slugify_field_label(label: str) -> str:
    """
    Match AssetField.save() behavior: slugify(label).replace("-", "_")
    but keep it local so imports are predictable even before AssetField.save runs.
    """
    from django.utils.text import slugify
    s = slugify(label).replace("-", "_")
    return s[:255]


def _get_dropdown_list(name_contains: str):
    qs = DropdownList.objects.filter(is_active=True)
    dl = qs.filter(name__icontains=name_contains).first()
    if dl:
        return dl
    return qs.filter(slug__icontains=name_contains.replace(" ", "-")).first()


class Command(BaseCommand):
    help = (
        "Import Equipment Optional Fields mapping from an Excel file.\n\n"
        "Expected columns:\n"
        "  A: Category\n"
        "  B: Equipment\n"
        "  C..: Optional fields (headers)\n\n"
        "Non-empty values in optional field columns become allowed dropdown values\n"
        "for that Equipment + Field.\n"
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            required=True,
            help="Path to the .xlsx file (absolute or relative to manage.py).",
        )
        parser.add_argument(
            "--sheet",
            default=None,
            help="Optional sheet name. If omitted, uses the first sheet.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Do not write changes; just show what would happen.",
        )
        parser.add_argument(
            "--create-missing-options",
            action="store_true",
            help="Create missing Category/Equipment DropdownOption records if not found.",
        )
        parser.add_argument(
            "--update-equipment-parent",
            action="store_true",
            default=True,
            help="Ensure Asset Equipment option parent is set to the Category option (default: on).",
        )
        parser.add_argument(
            "--no-update-equipment-parent",
            action="store_false",
            dest="update_equipment_parent",
            help="Do NOT update equipment parent to match category.",
        )
        parser.add_argument(
            "--prune-missing",
            action="store_true",
            help="Deactivate EquipmentOptionalField rows that are not present in the import.",
        )

    @transaction.atomic
    def handle(self, *args, **opts):
        path = opts["path"]
        sheet_name = opts["sheet"]
        dry_run = opts["dry_run"]
        create_missing = opts["create_missing_options"]
        update_parent = opts["update_equipment_parent"]
        prune_missing = opts["prune_missing"]

        categories_list = _get_dropdown_list("Asset Categories")
        equipment_list = _get_dropdown_list("Asset Equipment")

        if not categories_list:
            raise CommandError('Could not find DropdownList "Asset Categories".')
        if not equipment_list:
            raise CommandError('Could not find DropdownList "Asset Equipment".')

        try:
            wb = load_workbook(filename=path, data_only=True)
        except Exception as e:
            raise CommandError(f"Could not open workbook: {e}")

        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

        # Header row
        header_row = [c.value for c in ws[1]]
        if not header_row or len(header_row) < 3:
            raise CommandError("Header row is missing or does not contain optional field columns.")

        h0 = _norm(str(header_row[0] or ""))
        h1 = _norm(str(header_row[1] or ""))

        if h0 != "category" or h1 != "equipment":
            raise CommandError(
                f"Expected first two headers to be Category, Equipment. Got: {header_row[:2]}"
            )

        field_labels = [str(h or "").strip() for h in header_row[2:]]
        if not any(field_labels):
            raise CommandError("No optional field headers found after Category/Equipment.")

        # Build or fetch AssetField records for all headers
        field_by_index = {}
        created_fields = 0
        for idx, label in enumerate(field_labels, start=3):  # Excel columns start at 1; optional fields start at col 3
            label = (label or "").strip()
            if not label:
                continue
            slug = _slugify_field_label(label)
            af, created = AssetField.objects.get_or_create(
                slug=slug,
                defaults={"label": label, "is_active": True},
            )
            if not created:
                # Keep label aligned if it changed (best effort)
                if (af.label or "").strip() != label:
                    af.label = label
                    if not dry_run:
                        af.save(update_fields=["label"])
            else:
                created_fields += 1
                if dry_run:
                    # roll back later by forcing exception? easier: just report.
                    pass
            field_by_index[idx] = af

        # Preload existing options by normalized label
        category_opts = list(
            DropdownOption.objects.filter(dropdown_list=categories_list)
            .only("id", "label", "parent_id")
        )
        equipment_opts = list(
            DropdownOption.objects.filter(dropdown_list=equipment_list)
            .only("id", "label", "parent_id")
        )

        cat_by_label = defaultdict(list)
        for o in category_opts:
            cat_by_label[_norm(o.label)].append(o)

        eq_by_label = defaultdict(list)
        for o in equipment_opts:
            eq_by_label[_norm(o.label)].append(o)

        # Accumulate values per (equipment_id, field_id)
        values_map = defaultdict(lambda: defaultdict(set))

        missing_categories = set()
        missing_equipment = set()
        total_rows = 0
        used_rows = 0

        for r in range(2, ws.max_row + 1):
            total_rows += 1
            cat_raw = ws.cell(row=r, column=1).value
            eq_raw = ws.cell(row=r, column=2).value

            cat_label = str(cat_raw or "").strip()
            eq_label = str(eq_raw or "").strip()
            if not cat_label or not eq_label:
                continue

            cat_key = _norm(cat_label)
            eq_key = _norm(eq_label)

            cat_candidates = cat_by_label.get(cat_key, [])
            cat_opt = cat_candidates[0] if cat_candidates else None

            if not cat_opt and create_missing:
                cat_opt = DropdownOption(
                    dropdown_list=categories_list,
                    label=cat_label,
                    parent=None,
                    is_active=True,
                )
                if not dry_run:
                    cat_opt.save()
                cat_by_label[cat_key].append(cat_opt)
                created_fields += 0
            elif not cat_opt:
                missing_categories.add(cat_label)
                continue

            eq_candidates = eq_by_label.get(eq_key, [])
            eq_opt = None

            # Prefer equipment option already under this category
            if eq_candidates:
                for cand in eq_candidates:
                    if cand.parent_id == cat_opt.id:
                        eq_opt = cand
                        break
                if not eq_opt:
                    eq_opt = eq_candidates[0]

            if not eq_opt and create_missing:
                eq_opt = DropdownOption(
                    dropdown_list=equipment_list,
                    label=eq_label,
                    parent=cat_opt,
                    is_active=True,
                )
                if not dry_run:
                    eq_opt.save()
                eq_by_label[eq_key].append(eq_opt)
            elif not eq_opt:
                missing_equipment.add(eq_label)
                continue

            # Ensure parent is correct (equipment belongs to category)
            if update_parent and eq_opt.parent_id != cat_opt.id:
                eq_opt.parent_id = cat_opt.id
                if not dry_run:
                    eq_opt.save(update_fields=["parent"])
                # keep cached candidates updated
                for cand in eq_by_label.get(eq_key, []):
                    if cand.id == eq_opt.id:
                        cand.parent_id = cat_opt.id

            # Parse optional fields values in row
            any_value = False
            for col_idx, af in field_by_index.items():
                cell_val = ws.cell(row=r, column=col_idx).value
                if cell_val is None:
                    continue
                s = str(cell_val).strip()
                if not s:
                    continue
                any_value = True
                values_map[eq_opt.id][af.id].add(s)

            if any_value:
                used_rows += 1

        # Upsert EquipmentOptionalField rows
        upserts = 0
        created_rows = 0
        touched_pairs = set()

        # Preload existing rows for fast updates
        existing = {
            (e.equipment_id, e.field_id): e
            for e in EquipmentOptionalField.objects.all().only("id", "equipment_id", "field_id", "values", "is_active")
        }

        for equipment_id, fields_dict in values_map.items():
            for field_id, values_set in fields_dict.items():
                values_list = sorted(values_set, key=lambda x: _norm(x))
                key = (equipment_id, field_id)
                touched_pairs.add(key)

                row = existing.get(key)
                if row:
                    # update if changed or inactive
                    if (row.values or []) != values_list or not row.is_active:
                        row.values = values_list
                        row.is_active = True
                        upserts += 1
                        if not dry_run:
                            row.save(update_fields=["values", "is_active"])
                else:
                    created_rows += 1
                    if not dry_run:
                        EquipmentOptionalField.objects.create(
                            equipment_id=equipment_id,
                            field_id=field_id,
                            values=values_list,
                            is_active=True,
                        )

        pruned = 0
        if prune_missing:
            for key, row in existing.items():
                if key not in touched_pairs and row.is_active:
                    pruned += 1
                    if not dry_run:
                        row.is_active = False
                        row.save(update_fields=["is_active"])

        # Report
        self.stdout.write(self.style.SUCCESS("Import completed."))
        self.stdout.write(f"Workbook: {path}")
        self.stdout.write(f"Sheet: {ws.title}")
        self.stdout.write(f"Rows scanned: {total_rows}, rows used (had optional values): {used_rows}")
        self.stdout.write(f"AssetField created from headers: {created_fields}")
        self.stdout.write(f"EquipmentOptionalField created: {created_rows}, updated: {upserts}")
        if prune_missing:
            self.stdout.write(f"EquipmentOptionalField deactivated (pruned): {pruned}")

        if missing_categories:
            self.stdout.write(self.style.WARNING(f"Missing categories (skipped): {len(missing_categories)}"))
            for x in sorted(missing_categories)[:25]:
                self.stdout.write(f"  - {x}")
            if len(missing_categories) > 25:
                self.stdout.write("  ... (more)")

        if missing_equipment:
            self.stdout.write(self.style.WARNING(f"Missing equipment (skipped): {len(missing_equipment)}"))
            for x in sorted(missing_equipment)[:25]:
                self.stdout.write(f"  - {x}")
            if len(missing_equipment) > 25:
                self.stdout.write("  ... (more)")

        if dry_run:
            # force rollback of transaction
            raise CommandError("DRY RUN: Rolled back (no changes saved).")
