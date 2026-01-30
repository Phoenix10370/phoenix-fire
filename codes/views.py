# codes/views.py
import csv
from io import TextIOWrapper

import openpyxl

from django.contrib import messages
from django.http import JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.views.decorators.http import require_http_methods
from django.views.generic import CreateView, DeleteView, ListView, UpdateView

from .forms import AssetCodeForm, CodeForm, DefectCodeForm
from .models import (
    AssetCode,
    AssetField,
    Code,
    DefectCode,
    DropdownList,
    DropdownOption,
    EquipmentOptionalField,
)


# =========================
# EFSM CSV IMPORT
# =========================

@require_http_methods(["GET", "POST"])
def efsm_import(request):
    """
    Upload a CSV and import EFSM Codes.
    CSV headers required: code, fire_safety_measure, visits_per_year
    """
    if request.method == "POST":
        file = request.FILES.get("file")
        update_existing = request.POST.get("update_existing") == "on"

        if not file:
            messages.error(request, "Please choose a CSV file to upload.")
            return redirect("codes:efsm_import")

        name = (file.name or "").lower()
        if not name.endswith(".csv"):
            messages.error(request, "Only .csv files are supported.")
            return redirect("codes:efsm_import")

        created = 0
        updated = 0
        skipped = 0
        errors = []

        try:
            wrapper = TextIOWrapper(file.file, encoding="utf-8-sig")
            reader = csv.DictReader(wrapper)

            required = {"code", "fire_safety_measure", "visits_per_year"}
            if not reader.fieldnames or not required.issubset(set(reader.fieldnames)):
                messages.error(
                    request,
                    f"CSV must include headers: code, fire_safety_measure, visits_per_year. Found: {reader.fieldnames}",
                )
                return redirect("codes:efsm_import")

            for line_num, row in enumerate(reader, start=2):
                code_val = (row.get("code") or "").strip()
                fsm = (row.get("fire_safety_measure") or "").strip()
                vpy_raw = (row.get("visits_per_year") or "").strip()

                if not code_val or not fsm:
                    skipped += 1
                    errors.append(f"Row {line_num}: missing code or fire_safety_measure")
                    continue

                try:
                    vpy = int(vpy_raw) if vpy_raw else 1
                except ValueError:
                    skipped += 1
                    errors.append(f"Row {line_num}: invalid visits_per_year '{vpy_raw}'")
                    continue

                obj = Code.objects.filter(code=code_val).first()
                if obj:
                    if update_existing:
                        obj.fire_safety_measure = fsm
                        obj.visits_per_year = vpy
                        obj.save()
                        updated += 1
                    else:
                        skipped += 1
                else:
                    Code.objects.create(
                        code=code_val,
                        fire_safety_measure=fsm,
                        visits_per_year=vpy,
                    )
                    created += 1

        except Exception as e:
            messages.error(request, f"Import failed: {e}")
            return redirect("codes:efsm_import")

        messages.success(
            request,
            f"Import complete — created: {created}, updated: {updated}, skipped: {skipped}",
        )

        if errors:
            for msg in errors[:8]:
                messages.warning(request, msg)
            if len(errors) > 8:
                messages.warning(request, f"...and {len(errors) - 8} more row issues.")

        return redirect("codes:efsm_list")

    return render(
        request,
        "codes/efsm_import.html",
        {
            "title": "Import EFSM Codes",
            "cancel_url": reverse("codes:efsm_list"),
        },
    )


# =========================
# EFSM Codes
# =========================

class CodeListView(ListView):
    model = Code
    template_name = "codes/efsm_list.html"
    context_object_name = "items"


class CodeCreateView(CreateView):
    model = Code
    form_class = CodeForm
    template_name = "codes/add.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "Add EFSM Code"
        ctx["cancel_url"] = reverse_lazy("codes:efsm_list")
        return ctx

    def get_success_url(self):
        return reverse_lazy("codes:efsm_list")


class CodeUpdateView(UpdateView):
    model = Code
    form_class = CodeForm
    template_name = "codes/edit.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "Edit EFSM Code"
        ctx["cancel_url"] = reverse_lazy("codes:efsm_list")
        return ctx

    def get_success_url(self):
        return reverse_lazy("codes:efsm_list")


class CodeDeleteView(DeleteView):
    model = Code
    template_name = "codes/delete.html"
    success_url = reverse_lazy("codes:efsm_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "Delete EFSM Code"
        ctx["cancel_url"] = reverse_lazy("codes:efsm_list")
        return ctx


# =========================
# Defect Codes
# =========================

class DefectCodeListView(ListView):
    model = DefectCode
    template_name = "codes/defect_list.html"
    context_object_name = "items"


class DefectCodeCreateView(CreateView):
    model = DefectCode
    form_class = DefectCodeForm
    template_name = "codes/add.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "Add Defect Code"
        ctx["cancel_url"] = reverse_lazy("codes:defect_list")
        return ctx

    def get_success_url(self):
        return reverse_lazy("codes:defect_list")


class DefectCodeUpdateView(UpdateView):
    model = DefectCode
    form_class = DefectCodeForm
    template_name = "codes/edit.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "Edit Defect Code"
        ctx["cancel_url"] = reverse_lazy("codes:defect_list")
        return ctx

    def get_success_url(self):
        return reverse_lazy("codes:defect_list")


class DefectCodeDeleteView(DeleteView):
    model = DefectCode
    template_name = "codes/delete.html"
    success_url = reverse_lazy("codes:defect_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "Delete Defect Code"
        ctx["cancel_url"] = reverse_lazy("codes:defect_list")
        return ctx


# =========================
# Asset Codes (Dynamic Fields Support)
# =========================

class AssetDynamicFieldsMixin:
    """
    Adds dynamic equipment-based fields to the Asset form UI.
    Saves into AssetCode.attributes (JSON) without changing models/forms.

    Template gets:
      - dynamic_fields: list of dicts {slug,label,allowed_values,value}
    Posted values should be named:
      - attr_<field_slug>
    """

    def _get_equipment_optional_fields(self, equipment_id):
        if not equipment_id:
            return []
        return list(
            EquipmentOptionalField.objects.filter(
                equipment_id=equipment_id,
                is_active=True,
                field__is_active=True,
            )
            .select_related("field")
            .order_by("field__label")
        )

    def _build_dynamic_fields_context(self, equipment_id, attributes):
        attributes = attributes or {}
        if not isinstance(attributes, dict):
            attributes = {}

        rows = []
        for eof in self._get_equipment_optional_fields(equipment_id):
            slug = eof.field.slug
            rows.append(
                {
                    "slug": slug,
                    "label": eof.field.label,
                    "allowed_values": eof.values or [],
                    "value": attributes.get(slug, ""),
                }
            )
        return rows

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # Determine equipment id:
        equipment_id = None
        if self.request.method == "POST":
            equipment_id = self.request.POST.get("equipment") or None

        if (
            not equipment_id
            and getattr(self, "object", None) is not None
            and getattr(self.object, "equipment_id", None)
        ):
            equipment_id = self.object.equipment_id

        # Determine attributes to display:
        if self.request.method == "POST":
            # Echo back posted values so invalid form doesn't wipe them
            posted_attrs = {}
            for eof in self._get_equipment_optional_fields(equipment_id):
                key = f"attr_{eof.field.slug}"
                posted_attrs[eof.field.slug] = self.request.POST.get(key, "")
            attributes = posted_attrs
        else:
            attributes = {}
            if getattr(self, "object", None) is not None:
                attributes = self.object.attributes or {}

        ctx["dynamic_fields"] = self._build_dynamic_fields_context(equipment_id, attributes)
        return ctx

    def form_valid(self, form):
        """
        Save normal form fields, then update attributes from dynamic inputs.
        Only updates keys relevant to the asset's current equipment optional fields.
        """
        obj = form.save(commit=False)

        # Save core fields first so equipment is final
        obj.save()
        if hasattr(form, "save_m2m"):
            form.save_m2m()

        equipment_id = obj.equipment_id
        current_attrs = obj.attributes or {}
        if not isinstance(current_attrs, dict):
            current_attrs = {}

        # Update only the dynamic fields for this equipment
        for eof in self._get_equipment_optional_fields(equipment_id):
            slug = eof.field.slug
            post_key = f"attr_{slug}"

            # If not posted, skip (field not rendered)
            if post_key not in self.request.POST:
                continue

            raw_val = self.request.POST.get(post_key, "")
            val = raw_val.strip() if isinstance(raw_val, str) else raw_val

            # Remove key if empty
            if val is None or (isinstance(val, str) and val == ""):
                current_attrs.pop(slug, None)
            else:
                current_attrs[slug] = val

        obj.attributes = current_attrs
        obj.save(update_fields=["attributes"])

        self.object = obj
        return redirect(self.get_success_url())


# =========================
# Asset Codes
# =========================

class AssetCodeListView(ListView):
    model = AssetCode
    template_name = "codes/asset_list.html"
    context_object_name = "items"


class AssetCodeCreateView(AssetDynamicFieldsMixin, CreateView):
    model = AssetCode
    form_class = AssetCodeForm
    template_name = "codes/asset_form.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "Add Asset Code"
        ctx["cancel_url"] = reverse_lazy("codes:asset_list")
        return ctx

    def get_success_url(self):
        return reverse_lazy("codes:asset_list")


class AssetCodeUpdateView(AssetDynamicFieldsMixin, UpdateView):
    model = AssetCode
    form_class = AssetCodeForm
    template_name = "codes/asset_form.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "Edit Asset Code"
        ctx["cancel_url"] = reverse_lazy("codes:asset_list")
        return ctx

    def get_success_url(self):
        return reverse_lazy("codes:asset_list")


class AssetCodeDeleteView(DeleteView):
    model = AssetCode
    template_name = "codes/delete.html"
    success_url = reverse_lazy("codes:asset_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "Delete Asset Code"
        ctx["cancel_url"] = reverse_lazy("codes:asset_list")
        return ctx


@require_http_methods(["POST"])
def asset_bulk_delete(request):
    """
    Bulk delete AssetCode records selected on the Asset list page.
    Expects POST ids=1&ids=2...
    """
    raw_ids = request.POST.getlist("ids")

    if not raw_ids:
        messages.warning(request, "No Asset Codes selected.")
        return redirect("codes:asset_list")

    # Sanitize ids (ints only)
    ids = []
    for v in raw_ids:
        try:
            ids.append(int(v))
        except (TypeError, ValueError):
            continue

    if not ids:
        messages.warning(request, "No valid Asset Codes selected.")
        return redirect("codes:asset_list")

    qs = AssetCode.objects.filter(pk__in=ids)
    count = qs.count()

    if count == 0:
        messages.warning(request, "No matching Asset Codes found to delete.")
        return redirect("codes:asset_list")

    qs.delete()
    messages.success(request, f"Deleted {count} Asset Code(s).")
    return redirect("codes:asset_list")


# Optional: detail view (use if/when you add an asset detail template)
@require_http_methods(["GET"])
def asset_detail(request, pk):
    obj = get_object_or_404(AssetCode, pk=pk)
    # Precomputed "only fields with data" list (label/value)
    attribute_items = obj.get_attribute_items_display() if hasattr(obj, "get_attribute_items_display") else []
    return render(
        request,
        "codes/asset_detail.html",
        {
            "title": f"Asset Code {obj.code}",
            "object": obj,
            "attribute_items": attribute_items,
            "cancel_url": reverse("codes:asset_list"),
        },
    )


# =========================
# Asset Codes Excel Import (.xlsx)
# =========================

@require_http_methods(["GET", "POST"])
def asset_import(request):
    """
    Upload an Excel (.xlsx) file and import Asset Codes.

    Expected:
      - Row 1: headers
      - Column A header: "Main category" (required)
      - Column B header: "Equipment" (required)
      - Optional header: "Frequency" (if present, imported; else defaults to 1)

    Other columns become dynamic fields stored in AssetCode.attributes JSON.
    Only non-empty values are stored.
    """
    if request.method == "POST":
        file = request.FILES.get("file")
        update_existing = request.POST.get("update_existing") == "on"

        if not file:
            messages.error(request, "Please choose an Excel (.xlsx) file to upload.")
            return redirect("codes:asset_import")

        name = (file.name or "").lower()
        if not name.endswith(".xlsx"):
            messages.error(request, "Only .xlsx files are supported for Asset Codes import.")
            return redirect("codes:asset_import")

        created = 0
        updated = 0
        skipped = 0
        errors = []

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active

            # Read headers from row 1
            headers = []
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=1, column=col).value
                headers.append((str(v).strip() if v is not None else "").strip())

            def header_index(h):
                try:
                    return headers.index(h)
                except ValueError:
                    return None

            idx_cat = header_index("Main category")
            idx_eq = header_index("Equipment")
            idx_freq = header_index("Frequency")  # optional

            missing = []
            if idx_cat is None:
                missing.append("Main category")
            if idx_eq is None:
                missing.append("Equipment")

            if missing:
                messages.error(
                    request,
                    f"Excel Row 1 must include required headers: {', '.join(missing)}. Found: {headers}",
                )
                return redirect("codes:asset_import")

            cat_list = (
                DropdownList.objects.filter(slug="asset-categories", is_active=True).first()
                or DropdownList.objects.filter(slug="asset-category", is_active=True).first()
            )
            eq_list = DropdownList.objects.filter(slug="asset-equipment", is_active=True).first()

            if not cat_list or not eq_list:
                messages.error(
                    request,
                    "Missing required dropdown lists: 'Asset Categories' and/or 'Asset Equipment'. "
                    "Please confirm DropdownList slugs: asset-categories and asset-equipment.",
                )
                return redirect("codes:asset_import")

            required_headers = {"Main category", "Equipment"}
            for h in headers:
                if not h or h in required_headers:
                    continue
                AssetField.objects.update_or_create(
                    label=h,
                    defaults={"is_active": True},
                )

            field_slug_by_label = {f.label: f.slug for f in AssetField.objects.all()}

            for row_num in range(2, ws.max_row + 1):
                row_vals = []
                for col in range(1, ws.max_column + 1):
                    row_vals.append(ws.cell(row=row_num, column=col).value)

                raw_cat = row_vals[idx_cat]
                raw_eq = row_vals[idx_eq]

                if (raw_cat is None or str(raw_cat).strip() == "") and (raw_eq is None or str(raw_eq).strip() == ""):
                    skipped += 1
                    continue

                if raw_cat is None or str(raw_cat).strip() == "":
                    skipped += 1
                    errors.append(f"Row {row_num}: missing Main category")
                    continue
                if raw_eq is None or str(raw_eq).strip() == "":
                    skipped += 1
                    errors.append(f"Row {row_num}: missing Equipment")
                    continue

                cat_label = str(raw_cat).strip()
                eq_label = str(raw_eq).strip()

                freq = 1
                if idx_freq is not None:
                    raw_freq = row_vals[idx_freq]
                    try:
                        if raw_freq is not None and str(raw_freq).strip() != "":
                            freq = int(raw_freq)
                    except Exception:
                        errors.append(f"Row {row_num}: invalid Frequency '{raw_freq}', defaulted to 1")
                        freq = 1

                category_opt, _ = DropdownOption.objects.get_or_create(
                    dropdown_list=cat_list,
                    label=cat_label,
                    defaults={"is_active": True},
                )

                equipment_opt, created_eq = DropdownOption.objects.get_or_create(
                    dropdown_list=eq_list,
                    label=eq_label,
                    defaults={
                        "parent": category_opt,
                        "is_active": True,
                    },
                )

                if not created_eq and equipment_opt.parent_id != category_opt.id:
                    equipment_opt.parent = category_opt
                    equipment_opt.save(update_fields=["parent"])

                attributes = {}
                for idx, h in enumerate(headers):
                    if not h or h in required_headers or h == "Frequency":
                        continue
                    val = row_vals[idx]
                    if val is None:
                        continue
                    if isinstance(val, str) and not val.strip():
                        continue

                    slug = field_slug_by_label.get(h)
                    if not slug:
                        tmp = AssetField(label=h)
                        tmp.save()
                        slug = tmp.slug
                        field_slug_by_label[h] = slug

                    attributes[slug] = val

                if update_existing:
                    existing_qs = AssetCode.objects.filter(
                        category=category_opt,
                        equipment=equipment_opt,
                        frequency=freq,
                    ).order_by("created_at")

                    obj = existing_qs.first()
                    if obj:
                        obj.attributes = attributes
                        obj.is_active = True
                        obj.save()
                        updated += 1
                    else:
                        AssetCode.objects.create(
                            category=category_opt,
                            equipment=equipment_opt,
                            frequency=freq,
                            attributes=attributes,
                            is_active=True,
                        )
                        created += 1
                else:
                    AssetCode.objects.create(
                        category=category_opt,
                        equipment=equipment_opt,
                        frequency=freq,
                        attributes=attributes,
                        is_active=True,
                    )
                    created += 1

        except Exception as e:
            messages.error(request, f"Import failed: {e}")
            return redirect("codes:asset_import")

        messages.success(
            request,
            f"Import complete — created: {created}, updated: {updated}, skipped: {skipped}",
        )

        if errors:
            for msg in errors[:8]:
                messages.warning(request, msg)
            if len(errors) > 8:
                messages.warning(request, f"...and {len(errors) - 8} more row issues.")

        return redirect("codes:asset_list")

    return render(
        request,
        "codes/asset_import.html",
        {
            "title": "Import Asset Codes",
            "cancel_url": reverse("codes:asset_list"),
        },
    )


# =========================
# AJAX endpoint for dependent dropdown
# =========================

def asset_equipment_options(request):
    """
    Returns equipment options for a given category ID.
    Used by Asset Code form via fetch().
    """
    category_id = request.GET.get("category_id")
    eq_list = DropdownList.objects.filter(slug="asset-equipment", is_active=True).first()
    if not (category_id and eq_list):
        return JsonResponse({"results": []})

    qs = DropdownOption.objects.filter(
        dropdown_list=eq_list,
        is_active=True,
        parent_id=category_id,
    ).order_by("label")

    return JsonResponse(
        {"results": [{"id": o.id, "label": o.label} for o in qs]}
    )
